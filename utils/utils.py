import re

import pandas as pd
from django.core.exceptions import ValidationError
from django.core.files.storage import default_storage
from openpyxl.reader.excel import load_workbook

from product.models import FinalProduct, FinalPriceHistory, SellPriceHistory, PrimaryIngredient, PriceHistory, Unit


def get_last_final_price(product):
    if FinalPriceHistory.objects.filter(final_product=product, sell_price__gt=0).first():
        return FinalPriceHistory.objects.filter(final_product=product, sell_price__gt=0).first().sell_price


def get_last_sell_price(product):
    if SellPriceHistory.objects.filter(final_product=product, sell_price__gt=0).first():
        return SellPriceHistory.objects.filter(final_product=product, sell_price__gt=0).first().sell_price


def get_profit(product):
    final_price_history_qs = FinalPriceHistory.objects.filter(final_product=product, sell_price__gt=0)
    sell_price_history_qs = SellPriceHistory.objects.filter(final_product=product, sell_price__gt=0)
    if final_price_history_qs.exists() and sell_price_history_qs.exists():
        return final_price_history_qs.first().sell_price - sell_price_history_qs.first().sell_price


def get_last_price_history(primary_product: PrimaryIngredient):
    return PriceHistory.objects.filter(ingredient=primary_product).first().unit_price


def format_number_excel(x):
    try:
        if type(x) == float:
            return x
        return '{:,.0f}'.format(float(x))
    except (ValueError, TypeError):
        return x


def create_data(final_product: FinalProduct):
    data = pd.DataFrame({
        'نام محصول': [i.base_ingredient.name for i in final_product.ingredients.all()],
        'واحد': [i.base_ingredient.unit.title for i in final_product.ingredients.all()],
        'نسبت مورد نیاز': [i.unit_amount for i in final_product.ingredients.all()],
        'قیمت نهایی': [get_last_price_history(i.base_ingredient) for i in final_product.ingredients.all()]
    })

    # Format numbers with thousands separators
    data['نسبت مورد نیاز'] = data['نسبت مورد نیاز'].apply(format_number_excel)
    data['قیمت نهایی'] = data['قیمت نهایی'].apply(format_number_excel)

    empty_row = pd.DataFrame({
        'نام محصول': [' '],
        'واحد': [' '],
        'نسبت مورد نیاز': [' '],
        'قیمت نهایی': [' ']
    })

    titles_row = pd.DataFrame({
        'نام محصول': ['نام محصول نهایی'],
        'واحد': ['مجموع قیمت محاسبه شده'],
        'نسبت مورد نیاز': ['قیمت وارد شده در منو'],
        'قیمت نهایی': ['سود یا زیان']
    })

    additional_calculation = pd.DataFrame({
        'نام محصول': [final_product.name],
        'واحد': [get_last_sell_price(final_product)],
        'نسبت مورد نیاز': [get_last_final_price(final_product)],
        'قیمت نهایی': [get_profit(final_product)]
    })

    # Format numbers with thousands separators
    additional_calculation['واحد'] = additional_calculation['واحد'].apply(format_number_excel)
    additional_calculation['نسبت مورد نیاز'] = additional_calculation['نسبت مورد نیاز'].apply(format_number_excel)
    additional_calculation['قیمت نهایی'] = additional_calculation['قیمت نهایی'].apply(format_number_excel)

    data = pd.concat([data, empty_row, titles_row, additional_calculation], ignore_index=True)

    # HTML(style)

    return data


def import_from_excel(imported_file):
    try:
        wb = load_workbook(imported_file.path)
        ws = wb['Page 1']
        all_rows = list(ws.rows)
        for row in range(2, len(all_rows)):
            product_name = all_rows[row][1].value
            product_unit = all_rows[row][2].value
            product_unit_price = int(all_rows[row][3].value)
            if not PrimaryIngredient.objects.filter(name=product_name).exists():
                PrimaryIngredient.objects.create(name=product_name,
                                                 unit=Unit.objects.get_or_create(title=product_unit)[0])
            PriceHistory.objects.create(unit_price=product_unit_price,
                                        ingredient=PrimaryIngredient.objects.get(name=product_name))
    except ValidationError as v:
        raise ValidationError(v)
    except Exception as e:
        raise Exception('فایل اکسل وارد شده در فرمت درستی نمی باشد!')


def validate_excel(imported_file):
    try:
        name = default_storage.save(imported_file.name, imported_file)
        path = default_storage.path(name)

        wb = load_workbook(path)
        ws = wb['Page 1']
        all_rows = list(ws.rows)
        for row in range(2, len(all_rows)):
            product_name = all_rows[row][1].value
            product_unit = all_rows[row][2].value
            product_unit_price = int(all_rows[row][3].value)
        default_storage.delete(path)
    except Exception as e:
        try:
            default_storage.delete(path)
        except Exception:
            pass
        raise Exception('فایل اکسل وارد شده در فرمت درستی نمی باشد!')


def persian_to_english_number(persian_number):
    persian_to_english = {'۰': '0', '۱': '1', '۲': '2', '۳': '3', '۴': '4', '۵': '5', '۶': '6', '۷': '7', '۸': '8',
                          '۹': '9'}
    english_number = ''.join(
        persian_to_english[digit] if digit in persian_to_english else digit for digit in str(persian_number))
    return english_number


def format_number(number):
    num = persian_to_english_number(number)
    num_str = str(num)
    length = len(num_str)

    # Check if the length is not a multiple of 3
    if length % 3 != 0:
        # Calculate the number of leading zeros needed
        leading_zeros = 3 - (length % 3)
        # Add leading zeros to make the length a multiple of 3
        num_str = '0' * leading_zeros + num_str

    # Remove leading zeros
    num_str = num_str.lstrip('0')

    # Reverse the string, group the digits into sets of three, and join them with commas
    result = ','.join([num_str[::-1][i:i + 3] for i in range(0, len(num_str), 3)])

    return str(result[::-1]) + ' ریال '


def get_color(profit: int):
    if profit == 0:
        color = 'white'
    elif profit < 0:
        color = '#ce3d3a'
    else:
        color = 'green'
    return color


def denormalizer(text):
    text = text.replace('ی', 'ي')
    text = text.replace('ک', 'ك')
    text = text.strip()
    return text
